# Copyright (C) 2026  Edin Jelacic — AGPL-3.0-or-later
"""File converter for industrial documents into structured markdown.

Supported formats: PDF, XLSX, CSV, MD, TXT.

PDFs use PyMuPDF for fast text + image extraction and pdfplumber for accurate
table extraction. Images are described via an Ollama vision model when
VISION_MODEL is configured; otherwise their presence is noted as a placeholder.
"""

from __future__ import annotations

import argparse
import base64
import datetime
import os
import re
from pathlib import Path
from typing import Any, Callable

import pandas as pd

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

VISION_MODEL = os.environ.get("VISION_MODEL", "").strip()
OLLAMA_URL   = os.environ.get("OLLAMA_URL",   "http://localhost:11434")

# Skip images smaller than this (filters out icons / decorative elements)
_MIN_IMG_WIDTH  = 100
_MIN_IMG_HEIGHT = 100

# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

_DATE_RE     = re.compile(r"\b(\d{1,2}[/\-.]\d{1,2}[/\-.]\d{2,4})\b")
_REVISION_RE = re.compile(r"\b(v?\d+\.\d+(?:\.\d+)?)\b")


def _find_first_dates(text: str) -> list[datetime.date]:
    dates: list[datetime.date] = []
    for m in _DATE_RE.finditer(text):
        raw = m.group(1)
        for fmt in ("%m/%d/%Y", "%d-%m-%Y", "%m-%d-%Y", "%Y-%m-%d",
                    "%m/%d/%y", "%d-%m-%y"):
            try:
                dates.append(datetime.datetime.strptime(raw, fmt).date())
                break
            except ValueError:
                continue
    return sorted(dates)


def _find_first_revision(text: str) -> str | None:
    m = _REVISION_RE.search(text)
    return m.group(0) if m else None


def _escape_md_cell(val: Any) -> str:
    return str(val).replace("|", "\\|").replace("\n", " ").strip()


def _dataframe_to_md(df: pd.DataFrame) -> str:
    headers = df.columns.astype(str).tolist()
    lines   = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    for _, row in df.iterrows():
        lines.append("| " + " | ".join(_escape_md_cell(v) for v in row.values) + " |")
    return "\n".join(lines)


def _plumber_table_to_md(table: list[list]) -> str:
    """Convert a pdfplumber table (list of rows, each row a list of cells) to markdown."""
    if not table or not table[0]:
        return ""
    header = [_escape_md_cell(c) for c in table[0]]
    lines  = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join("---" for _ in header) + " |",
    ]
    for row in table[1:]:
        cells = [_escape_md_cell(c) for c in row]
        cells += [""] * max(0, len(header) - len(cells))
        lines.append("| " + " | ".join(cells[: len(header)]) + " |")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Vision model image description
# ---------------------------------------------------------------------------


def _describe_image(img_bytes: bytes) -> str | None:
    """Describe an image using the configured Ollama vision model."""
    if not VISION_MODEL:
        return None
    try:
        import ollama
        resp = ollama.chat(
            model=VISION_MODEL,
            messages=[
                {
                    "role": "user",
                    "content": (
                        "Describe this technical diagram, chart, or image concisely. "
                        "Focus on labels, measurements, data values, and what it depicts."
                    ),
                    "images": [base64.b64encode(img_bytes).decode()],
                }
            ],
        )
        return resp["message"]["content"].strip()
    except Exception:
        return None


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------


def _convert_pdf(
    file_path: str,
    progress_callback: Callable[[str], None] | None = None,
) -> tuple[str, dict[str, Any]]:
    """Extract structured markdown from a PDF.

    Uses PyMuPDF for fast text + image extraction, pdfplumber for tables.
    """
    try:
        import fitz  # PyMuPDF
    except ImportError:
        raise ImportError("Install PyMuPDF: pip install pymupdf")

    import pdfplumber

    path  = Path(file_path)
    title = path.stem

    doc        = fitz.open(file_path)
    page_count = len(doc)
    md_parts: list[str] = []

    with pdfplumber.open(file_path) as pdf:
        for page_idx in range(page_count):
            if progress_callback:
                progress_callback(f"Converting page {page_idx + 1}/{page_count} …")

            fitz_page  = doc[page_idx]
            plumb_page = pdf.pages[page_idx]

            heading = (
                f"## Page {page_idx + 1}" if page_count > 1 else f"# {title}"
            )
            section: list[str] = [heading]

            # Text via PyMuPDF — significantly faster than pdfplumber for plain text
            text = fitz_page.get_text("text") or ""
            if text.strip():
                section.append(f"### Text\n\n{text.strip()}")

            # Tables via pdfplumber — more accurate for complex layouts
            tables = plumb_page.extract_tables()
            if tables:
                valid = [(i, t) for i, t in enumerate(tables, 1) if t]
                for t_idx, table in valid:
                    md = _plumber_table_to_md(table)
                    if md:
                        label = (
                            f"### Table {t_idx}" if len(valid) > 1 else "### Table"
                        )
                        section.append(f"{label}\n\n{md}")

            # Images via PyMuPDF
            img_refs  = fitz_page.get_images(full=True)
            img_parts: list[str] = []
            for img_num, img_ref in enumerate(img_refs, start=1):
                xref = img_ref[0]
                try:
                    info = doc.extract_image(xref)
                    w, h = info.get("width", 0), info.get("height", 0)
                    if w < _MIN_IMG_WIDTH or h < _MIN_IMG_HEIGHT:
                        continue  # skip tiny decorative images
                    desc = _describe_image(info["image"])
                    if desc:
                        img_parts.append(f"**[Figure {img_num}]** ({w}×{h}px): {desc}")
                    else:
                        hint = (
                            " Set VISION_MODEL env var for automatic descriptions."
                            if not VISION_MODEL
                            else ""
                        )
                        img_parts.append(
                            f"**[Figure {img_num}]** ({w}×{h}px): "
                            f"*Technical diagram/image present.*{hint}"
                        )
                except Exception:
                    continue

            if img_parts:
                section.append("### Figures\n\n" + "\n\n".join(img_parts))

            md_parts.append("\n\n".join(section))
            md_parts.append("---")

    doc.close()

    full_md     = "\n\n".join(md_parts)
    first_dates = _find_first_dates(full_md)

    return full_md, {
        "title":             title,
        "detected_date":     first_dates[0].isoformat() if first_dates else None,
        "detected_revision": _find_first_revision(full_md),
        "page_count":        page_count,
    }


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------


def _convert_xlsx(
    file_path: str,
    progress_callback: Callable[[str], None] | None = None,
) -> tuple[str, dict[str, Any]]:
    """Convert XLSX workbook to structured markdown.

    Each sheet becomes a ## Sheet: <name> section.
    Merged cells are handled by column-wise fill-down.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("Install openpyxl: pip install openpyxl")

    wb           = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    md_parts:    list[str] = []
    sheet_names: list[str] = []
    total        = len(wb.sheetnames)

    for sheet_idx, sheet_name in enumerate(wb.sheetnames, start=1):
        if progress_callback:
            progress_callback(
                f"Converting sheet {sheet_idx}/{total}: {sheet_name} …"
            )

        ws = wb[sheet_name]
        sheet_names.append(sheet_name)

        num_rows = ws.max_row or 0
        num_cols = ws.max_column or 0
        if num_rows == 0 or num_cols == 0:
            continue

        raw: dict[tuple[int, int], Any] = {}
        for row in ws.iter_rows():
            for cell in row:
                raw[(cell.row - 1, cell.column - 1)] = cell.value

        fill_down: dict[int, Any] = {}
        for col in range(num_cols):
            for row in range(num_rows):
                val = raw.get((row, col))
                if val is not None:
                    fill_down[col] = val
                elif col in fill_down:
                    raw[(row, col)] = fill_down[col]

        rows_list = [
            tuple(raw.get((row, col)) for col in range(num_cols))
            for row in range(num_rows)
            if any(raw.get((row, col)) is not None for col in range(num_cols))
        ]
        if not rows_list:
            continue

        md_parts.append(f"## Sheet: {sheet_name}")
        header = [_escape_md_cell(c) for c in rows_list[0]]
        md_parts.append("| " + " | ".join(header) + " |")
        md_parts.append("| " + " | ".join("---" for _ in header) + " |")
        for data_row in rows_list[1:]:
            cells = [_escape_md_cell(c) for c in list(data_row)[: len(header)]]
            cells += [""] * (len(header) - len(cells))
            md_parts.append("| " + " | ".join(cells) + " |")
        md_parts.append("")

    wb.close()
    text        = "\n".join(md_parts)
    first_dates = _find_first_dates(text)

    return text, {
        "title":             Path(file_path).stem,
        "detected_date":     first_dates[0].isoformat() if first_dates else None,
        "detected_revision": _find_first_revision(text),
        "sheet_names":       sheet_names,
    }


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------


def _convert_csv(file_path: str) -> tuple[str, dict[str, Any]]:
    """Convert CSV to markdown table, splitting large files into chunks."""
    df         = pd.read_csv(file_path)
    chunk_size = 100
    md_parts:  list[str] = []

    if len(df) <= chunk_size:
        md_parts.append(f"# {Path(file_path).stem}")
        md_parts.append(_dataframe_to_md(df))
    else:
        n = (len(df) + chunk_size - 1) // chunk_size
        for i in range(n):
            start, end = i * chunk_size, min((i + 1) * chunk_size, len(df))
            md_parts.append(f"## Chunk {i + 1}/{n} (rows {start + 1}–{end})")
            md_parts.append(_dataframe_to_md(df.iloc[start:end]))
            md_parts.append("")

    text        = "\n\n".join(md_parts)
    first_dates = _find_first_dates(text)

    return text, {
        "title":             Path(file_path).stem,
        "detected_date":     first_dates[0].isoformat() if first_dates else None,
        "detected_revision": _find_first_revision(text),
        "row_count":         len(df),
        "column_count":      len(df.columns),
    }


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def convert_file(
    file_path: str,
    progress_callback: Callable[[str], None] | None = None,
) -> tuple[str, dict[str, Any]]:
    """Convert a file to structured markdown suitable for chunking.

    Parameters
    ----------
    file_path         : Path to the file (.pdf, .xlsx, .csv, .md, .txt).
    progress_callback : Optional callable(str) for per-page/sheet updates.

    Returns
    -------
    tuple[str, dict]
        (markdown_string, metadata_dict). Metadata always includes
        *title*, *detected_date*, *detected_revision*.
    """
    path = Path(file_path)
    ext  = path.suffix.lower()

    if ext == ".pdf":
        return _convert_pdf(file_path, progress_callback=progress_callback)
    elif ext in (".xlsx", ".xls"):
        return _convert_xlsx(file_path, progress_callback=progress_callback)
    elif ext == ".csv":
        return _convert_csv(file_path)
    elif ext in (".md", ".txt"):
        content = path.read_text(encoding="utf-8", errors="replace")
        return content, {
            "title":             path.stem,
            "detected_date":     None,
            "detected_revision": None,
            "file_type":         ext.lstrip("."),
        }
    else:
        raise ValueError(f"Unsupported file type: {ext}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Convert industrial documents to structured markdown"
    )
    parser.add_argument("--file", required=True, help="Path to the input file")
    args = parser.parse_args()

    import json

    markdown, metadata = convert_file(
        args.file, progress_callback=lambda m: print(f"  {m}")
    )
    print("---METADATA---")
    print(json.dumps(metadata, indent=2, default=str))
    print("---MARKDOWN---")
    print(markdown)


if __name__ == "__main__":
    main()
